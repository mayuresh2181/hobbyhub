# NSE Delivery Breakout Scanner - Full NSE EQ Universe
# Fixed version: numeric dtype coercion now applied on BOTH the cache-hit
# and fresh-fetch paths (previously only fresh-fetch), plus min-history-days
# and divide-by-zero aguards.

import requests
import datetime as dt
import pandas as pd
import io
import time
import os
import concurrent.futures

from openpyxl import load_workbook
from openpyxl.styles import Font
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

BASE_FOLDER = "data/nse/delivery"
LOG_FOLDER = os.path.join(BASE_FOLDER, "daily_logs")
OUTPUT_FOLDER = os.path.join(BASE_FOLDER, "daily_output")

TRADING_DAYS_LOOKBACK = 30
MIN_DELIVERY_RATIO = 1
MIN_DELIVERY_VALUE_CR = 20
MIN_HISTORY_DAYS = 15   # require at least this many valid historical days before trusting the average
MAX_WORKERS = 5

NUMERIC_COLS = ["DELIV_QTY", "TTL_TRD_QNTY", "CLOSE_PRICE"]

os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def create_session():
    s = requests.Session()
    retry = Retry(total=5, backoff_factor=1,
                  status_forcelist=[429,500,502,503,504])
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    s.headers.update({
        "User-Agent":"Mozilla/5.0",
        "Referer":"https://www.nseindia.com/",
        "Accept-Language":"en-US,en;q=0.9"
    })
    s.get("https://www.nseindia.com", timeout=20)
    return s

def get_last_days(n):
    days=[]
    d=dt.date.today()-dt.timedelta(days=1)
    while len(days)<n:
        if d.weekday()<5:
            days.append(d)
        d-=dt.timedelta(days=1)
    return sorted(days)

def _coerce_numeric(df):
    """Force numeric dtype on the columns we do arithmetic on, no matter
    where the dataframe came from (cache or fresh fetch). This is the
    single guard that prevents object-dtype columns from ever reaching
    the division/round() calls downstream."""
    for c in NUMERIC_COLS:
        df[c] = pd.to_numeric(
            df[c].astype(str).str.replace(",", "").str.strip(),
            errors="coerce"
        )
    return df

def get_delivery_data(day, session):

    print(f"\nProcessing {day}")

    fp = os.path.join(LOG_FOLDER, f"{day:%Y%m%d}.csv")

    # =====================================================
    # CACHE EXISTS
    # =====================================================

    if os.path.exists(fp):

        print(f"Using cache: {fp}")

        try:
            df = pd.read_csv(fp)

            print("Cache Shape :", df.shape)

            if df.empty:
                print("Empty cache. Deleting...")
                os.remove(fp)
                return None

            df["DATE"] = pd.to_datetime(df["DATE"]).dt.date

            df = _coerce_numeric(df)

        except Exception as e:
            print(f"Cache read failed : {e}")

            if os.path.exists(fp):
                os.remove(fp)

            return None

    # =====================================================
    # DOWNLOAD FROM NSE
    # =====================================================

    else:

        print(f"Downloading: {day}")

        url = (
            "https://archives.nseindia.com/products/content/"
            f"sec_bhavdata_full_{day:%d%m%Y}.csv"
        )

        r = session.get(url, timeout=30)

        if r.status_code != 200:
            print(f"HTTP Error : {r.status_code}")
            return None

        if "SYMBOL" not in r.text:
            print("Invalid bhavcopy")
            return None

        df = pd.read_csv(io.StringIO(r.text))

        df.columns = (
            df.columns
            .str.strip()
            .str.upper()
            .str.replace(" ", "_")
        )

        print("Downloaded Shape :", df.shape)

        df["SERIES"] = df["SERIES"].astype(str).str.strip()

        print("Rows before EQ filter :", len(df))

        df = df[df["SERIES"] == "EQ"].copy()

        print("Rows after EQ filter :", len(df))

        df["SYMBOL"] = df["SYMBOL"].astype(str).str.strip()

        df["DATE"] = day

        df = _coerce_numeric(df)

    # =====================================================
    # COMMON VALIDATION
    # =====================================================

    df = df.dropna(subset=NUMERIC_COLS)

    print("Rows after dropna :", len(df))

    if df.empty:
        print("No valid rows")

        if os.path.exists(fp):
            os.remove(fp)

        return None

    # =====================================================
    # SAVE ONLY IF NEW DOWNLOAD
    # =====================================================

    if not os.path.exists(fp):
        df.to_csv(fp, index=False)

    return df

def download_all(days,session):
    out=[]
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs=[ex.submit(get_delivery_data,d,session) for d in days]
        for f in concurrent.futures.as_completed(futs):
            x=f.result()
            if x is not None:
                out.append(x)
    return out

def format_excel(path):
    wb=load_workbook(path)
    ws=wb.active
    ws.freeze_panes="A2"
    for c in ws[1]:
        c.font=Font(bold=True)
    for col in ws.columns:
        w=min(max(len(str(x.value)) if x.value else 0 for x in col)+3,35)
        ws.column_dimensions[col[0].column_letter].width=w
    wb.save(path)

def main():
    s=create_session()
    days=get_last_days(TRADING_DAYS_LOOKBACK)
    data=download_all(days,s)
    print("Number of DataFrames:", len(data))

    for d in data:
        print(d.shape)
        
    if not data:
        print("No data");return
    df=pd.concat(data,ignore_index=True)
    latest=max(df.DATE.unique())
    latest_df=df[df.DATE==latest].copy()
    hist=df[df.DATE<latest].copy()

    # average delivery qty AND a count of how many valid days back it up
    grp = hist.groupby("SYMBOL")["DELIV_QTY"]
    avg = grp.mean().reset_index(name="AVG_30_DELIVERY")
    counts = grp.count().reset_index(name="N_HIST_DAYS")
    avg = avg.merge(counts, on="SYMBOL")

    # guard: require a real trailing history, and a non-zero baseline
    # (a zero/near-zero average turns the ratio into inf or noise)
    avg = avg[(avg.N_HIST_DAYS >= MIN_HISTORY_DAYS) & (avg.AVG_30_DELIVERY > 0)]

    merged=latest_df.merge(avg,on="SYMBOL",how="inner")
    merged["DELIVERY_RATIO"]=(merged.DELIV_QTY/merged.AVG_30_DELIVERY).round(2)
    merged["DELIVERY_PERCENT"]=(merged.DELIV_QTY/merged.TTL_TRD_QNTY*100).round(2)
    merged["TRADED_VALUE_CR"]=(merged.TTL_TRD_QNTY*merged.CLOSE_PRICE/1e7).round(2)
    merged["DELIVERY_VALUE_CR"]=(merged.DELIV_QTY*merged.CLOSE_PRICE/1e7).round(2)

    res=merged[(merged.DELIVERY_RATIO>=MIN_DELIVERY_RATIO)&
               (merged.DELIVERY_VALUE_CR>=MIN_DELIVERY_VALUE_CR)].copy()

    res=res.sort_values(["DELIVERY_VALUE_CR","DELIVERY_RATIO"],ascending=False)

    res["SYMBOL"]=res["SYMBOL"].apply(
        lambda x:f'=HYPERLINK("https://www.tradingview.com/chart/?symbol=NSE:{x}","{x}")'
    )

    out=res[[
        "SYMBOL","CLOSE_PRICE","TTL_TRD_QNTY","DELIV_QTY",
        "DELIVERY_PERCENT","AVG_30_DELIVERY","DELIVERY_RATIO",
        "TRADED_VALUE_CR","DELIVERY_VALUE_CR"
    ]].copy()

    out.columns=[
        "SYMBOL","LTP","TOTAL_VOLUME","DELIVERY_QTY",
        "DELIVERY_%","AVG_30_DELIVERY","DELIVERY_RATIO",
        "TRADED_VALUE_CR","DELIVERY_VALUE_CR"
    ]

    outfile=os.path.join(OUTPUT_FOLDER,f"delivery_breakout_{latest:%Y%m%d}.xlsx")
    out.to_excel(outfile,index=False)
    format_excel(outfile)
    print(outfile)
    print(out.head(20))

    print(len(latest_df))
    print(len(hist))
    print(len(avg))
    print(len(merged))
    print(merged["DELIVERY_RATIO"].describe())

if __name__=="__main__":
    st=time.time()
    main()
    print("Execution:",round(time.time()-st,2),"sec")
