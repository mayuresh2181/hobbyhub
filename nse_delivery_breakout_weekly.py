# NSE Delivery Breakout Scanner - WEEKLY edition (full NSE EQ universe)
#
# What changed vs the daily version:
#   * Bhavcopies are still fetched per day, but aggregated into ISO weeks
#     (Mon-Fri, labelled by the Friday) before any ratio is computed.
#   * DELIVERY_RATIO = latest week's delivery vs the mean of the prior
#     HISTORY_WEEKS weeks (median ratio also reported as a robustness check).
#   * Partial weeks are handled by normalising to a per-trading-day basis,
#     so running mid-week (or after a holiday-shortened week) does not
#     mechanically depress the ratio.
#   * Per-thread HTTP sessions, negative caching for holidays/missing files,
#     archive host fallback, and a price-change column to spot corporate actions.

import requests
import datetime as dt
import pandas as pd
import io
import time
import os
import threading
import concurrent.futures

from openpyxl import load_workbook
from openpyxl.styles import Font
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

BASE_FOLDER = "data/nse/delivery"
LOG_FOLDER = os.path.join(BASE_FOLDER, "daily_logs")
OUTPUT_FOLDER = os.path.join(BASE_FOLDER, "weekly_output")

# ~15 calendar weeks of weekdays. Holidays get skipped automatically, so
# fetch a little more than HISTORY_WEEKS + 1 weeks' worth.
TRADING_DAYS_LOOKBACK = 75

HISTORY_WEEKS = 12          # prior weeks that form the baseline
MIN_HISTORY_WEEKS = 8       # need at least this many valid weeks to trust it
MIN_DAYS_PER_WEEK = 3       # a historical week with fewer valid days is dropped
MIN_DAYS_LATEST_WEEK = 2    # do not score the current week before this many days

# NOTE: weekly sums smooth out single-day spikes, so a 5x weekly surge is far
# rarer than a 5x daily surge. 2.5-3x is the equivalent screen tightness.
MIN_DELIVERY_RATIO = 2.5

# This is now a WEEKLY delivery value, i.e. roughly 5 daily bars added up.
# The old 50 Cr daily floor corresponds to ~250 Cr weekly.
MIN_DELIVERY_VALUE_CR = 200

# True  -> compare avg delivery per trading day (safe for partial/short weeks)
# False -> compare raw weekly totals (only run on completed weeks)
NORMALIZE_BY_DAYS = True

MAX_WORKERS = 5

NUMERIC_COLS = ["DELIV_QTY", "TTL_TRD_QNTY", "CLOSE_PRICE"]

ARCHIVE_HOSTS = [
    "https://nsearchives.nseindia.com",
    "https://archives.nseindia.com",
]

os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------

_local = threading.local()


def create_session():
    s = requests.Session()
    retry = Retry(total=5, backoff_factor=1,
                  status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    s.headers.update({
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://www.nseindia.com/",
        "Accept-Language": "en-US,en;q=0.9",
    })
    try:
        s.get("https://www.nseindia.com", timeout=20)   # warm the cookie jar
    except requests.RequestException:
        pass
    return s


def get_session():
    """requests.Session is not thread-safe; give every worker its own."""
    if not hasattr(_local, "session"):
        _local.session = create_session()
    return _local.session


# ---------------------------------------------------------------------------
# DATE HELPERS
# ---------------------------------------------------------------------------

def get_last_days(n, include_today=None):
    """Last n weekdays. Today is excluded before the bhavcopy publishes
    (roughly 18:30 IST) so we don't burn a slot on a guaranteed 404."""
    if include_today is None:
        include_today = dt.datetime.now().hour >= 19

    days = []
    d = dt.date.today()
    if not include_today:
        d -= dt.timedelta(days=1)

    while len(days) < n:
        if d.weekday() < 5:
            days.append(d)
        d -= dt.timedelta(days=1)
    return sorted(days)


def _coerce_numeric(df):
    """Force numeric dtype on the columns we do arithmetic on, no matter
    where the dataframe came from (cache or fresh fetch)."""
    for c in NUMERIC_COLS:
        df[c] = pd.to_numeric(
            df[c].astype(str).str.replace(",", "").str.strip(),
            errors="coerce",
        )
    return df


# ---------------------------------------------------------------------------
# FETCH
# ---------------------------------------------------------------------------

def _download_bhavcopy(day, session):
    """Try each archive host. Returns response text or None."""
    path = f"/products/content/sec_bhavdata_full_{day:%d%m%Y}.csv"
    for host in ARCHIVE_HOSTS:
        try:
            r = session.get(host + path, timeout=30)
        except requests.RequestException as e:
            print(f"  {host}: request failed ({e})")
            continue
        if r.status_code == 200 and "SYMBOL" in r.text:
            return r.text
        print(f"  {host}: HTTP {r.status_code}")
    return None


def get_delivery_data(day):
    session = get_session()
    fp = os.path.join(LOG_FOLDER, f"{day:%Y%m%d}.csv")
    missing_marker = os.path.join(LOG_FOLDER, f"{day:%Y%m%d}.missing")

    # A day we already know has no bhavcopy (holiday). Don't re-request it.
    if os.path.exists(missing_marker):
        return None

    # ---------------- cache ----------------
    if os.path.exists(fp):
        try:
            df = pd.read_csv(fp)
            if df.empty:
                os.remove(fp)
                return None
            df["DATE"] = pd.to_datetime(df["DATE"]).dt.date
            df = _coerce_numeric(df)
        except Exception as e:
            print(f"{day}: cache read failed ({e}) - deleting")
            try:
                os.remove(fp)
            except OSError:
                pass
            return None
        fresh = False

    # ---------------- download ----------------
    else:
        print(f"{day}: downloading")
        text = _download_bhavcopy(day, session)
        if text is None:
            # Holiday or not yet published. Mark it so future runs skip it,
            # but never mark the last few days (file may still appear).
            if (dt.date.today() - day).days > 5:
                open(missing_marker, "w").close()
            return None

        df = pd.read_csv(io.StringIO(text))
        df.columns = (df.columns.str.strip().str.upper().str.replace(" ", "_"))
        df["SERIES"] = df["SERIES"].astype(str).str.strip()
        df = df[df["SERIES"] == "EQ"].copy()
        df["SYMBOL"] = df["SYMBOL"].astype(str).str.strip()
        df["DATE"] = day
        df = _coerce_numeric(df)
        fresh = True

    # ---------------- common validation ----------------
    keep = ["SYMBOL", "DATE"] + NUMERIC_COLS
    df = df[[c for c in keep if c in df.columns]].dropna(subset=NUMERIC_COLS)

    if df.empty:
        print(f"{day}: no valid rows")
        if os.path.exists(fp):
            os.remove(fp)
        return None

    if fresh:
        df.to_csv(fp, index=False)

    return df


def download_all(days):
    out = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = [ex.submit(get_delivery_data, d) for d in days]
        for f in concurrent.futures.as_completed(futs):
            try:
                x = f.result()
            except Exception as e:
                print(f"worker failed: {e}")
                continue
            if x is not None:
                out.append(x)
    return out


# ---------------------------------------------------------------------------
# WEEKLY AGGREGATION
# ---------------------------------------------------------------------------

def build_weekly(df):
    """Collapse daily bhavcopy rows into one row per SYMBOL per week."""
    df = df.copy()
    df["DATE"] = pd.to_datetime(df["DATE"])
    df = df.sort_values("DATE")

    # Week labelled by its Friday. Mon-Fri of the same trading week group together.
    df["WEEK"] = df["DATE"].dt.to_period("W-FRI")

    # Value is summed day-by-day rather than qty x last close, so a big move
    # inside the week doesn't distort the turnover figure.
    df["DAY_TRD_VALUE"] = df["TTL_TRD_QNTY"] * df["CLOSE_PRICE"]
    df["DAY_DEL_VALUE"] = df["DELIV_QTY"] * df["CLOSE_PRICE"]

    wk = (
        df.groupby(["SYMBOL", "WEEK"], observed=True)
          .agg(
              WEEK_DELIV_QTY=("DELIV_QTY", "sum"),
              WEEK_TRD_QTY=("TTL_TRD_QNTY", "sum"),
              WEEK_TRD_VALUE=("DAY_TRD_VALUE", "sum"),
              WEEK_DEL_VALUE=("DAY_DEL_VALUE", "sum"),
              CLOSE_PRICE=("CLOSE_PRICE", "last"),
              DAYS_IN_WEEK=("DATE", "nunique"),
              WEEK_ENDING=("DATE", "max"),
          )
          .reset_index()
          .sort_values(["SYMBOL", "WEEK"])
    )

    # Week-on-week price change: a ~-50% / -80% print alongside a delivery
    # explosion is almost always a split or bonus, not accumulation.
    wk["PREV_CLOSE"] = wk.groupby("SYMBOL")["CLOSE_PRICE"].shift(1)

    return wk


def main():
    days = get_last_days(TRADING_DAYS_LOOKBACK)
    print(f"Requesting {len(days)} weekdays: {days[0]} -> {days[-1]}")

    data = download_all(days)
    print("Trading days retrieved:", len(data))
    if not data:
        print("No data")
        return

    df = pd.concat(data, ignore_index=True)
    wk = build_weekly(df)

    latest_week = wk["WEEK"].max()
    cur = wk[wk["WEEK"] == latest_week].copy()
    hist = wk[wk["WEEK"] < latest_week].copy()

    # Keep only the most recent HISTORY_WEEKS complete-enough weeks.
    hist = hist[hist["DAYS_IN_WEEK"] >= MIN_DAYS_PER_WEEK]
    keep_weeks = sorted(hist["WEEK"].unique())[-HISTORY_WEEKS:]
    hist = hist[hist["WEEK"].isin(keep_weeks)]

    print(f"Latest week : {latest_week} (ending {cur['WEEK_ENDING'].max().date()})")
    print(f"Baseline    : {len(keep_weeks)} weeks, {keep_weeks[0]} -> {keep_weeks[-1]}")

    # Metric being compared. Per-day normalisation keeps a 3-day holiday week
    # or a mid-week run from looking like a collapse in delivery.
    if NORMALIZE_BY_DAYS:
        hist["METRIC"] = hist["WEEK_DELIV_QTY"] / hist["DAYS_IN_WEEK"]
        cur["METRIC"] = cur["WEEK_DELIV_QTY"] / cur["DAYS_IN_WEEK"]
    else:
        hist["METRIC"] = hist["WEEK_DELIV_QTY"]
        cur["METRIC"] = cur["WEEK_DELIV_QTY"]

    grp = hist.groupby("SYMBOL")["METRIC"]
    base = pd.DataFrame({
        "AVG_WEEKLY_DELIVERY": grp.mean(),
        "MED_WEEKLY_DELIVERY": grp.median(),
        "N_HIST_WEEKS": grp.count(),
    }).reset_index()

    base = base[(base["N_HIST_WEEKS"] >= MIN_HISTORY_WEEKS) &
                (base["AVG_WEEKLY_DELIVERY"] > 0) &
                (base["MED_WEEKLY_DELIVERY"] > 0)]

    cur = cur[cur["DAYS_IN_WEEK"] >= MIN_DAYS_LATEST_WEEK]

    merged = cur.merge(base, on="SYMBOL", how="inner")
    if merged.empty:
        print("Nothing survived the history filters")
        return

    merged["DELIVERY_RATIO"] = (merged["METRIC"] / merged["AVG_WEEKLY_DELIVERY"]).round(2)
    merged["DELIVERY_RATIO_MED"] = (merged["METRIC"] / merged["MED_WEEKLY_DELIVERY"]).round(2)
    merged["DELIVERY_PERCENT"] = (merged["WEEK_DELIV_QTY"] / merged["WEEK_TRD_QTY"] * 100).round(2)
    merged["TRADED_VALUE_CR"] = (merged["WEEK_TRD_VALUE"] / 1e7).round(2)
    merged["DELIVERY_VALUE_CR"] = (merged["WEEK_DEL_VALUE"] / 1e7).round(2)
    merged["PRICE_CHG_PCT"] = (
        (merged["CLOSE_PRICE"] / merged["PREV_CLOSE"] - 1) * 100
    ).round(2)

    res = merged[(merged["DELIVERY_RATIO"] >= MIN_DELIVERY_RATIO) &
                 (merged["DELIVERY_VALUE_CR"] >= MIN_DELIVERY_VALUE_CR)].copy()

    if res.empty:
        print("No breakouts at current thresholds")
        print(merged["DELIVERY_RATIO"].describe())
        return

    res = res.sort_values(["DELIVERY_VALUE_CR", "DELIVERY_RATIO"], ascending=False)

    res["WEEK_ENDING"] = pd.to_datetime(res["WEEK_ENDING"]).dt.strftime("%Y-%m-%d")
    res["SYMBOL"] = res["SYMBOL"].apply(
        lambda x: f'=HYPERLINK("https://www.tradingview.com/chart/?symbol=NSE:{x}","{x}")'
    )

    out = res[[
        "SYMBOL", "WEEK_ENDING", "DAYS_IN_WEEK", "CLOSE_PRICE", "PRICE_CHG_PCT",
        "WEEK_TRD_QTY", "WEEK_DELIV_QTY", "DELIVERY_PERCENT",
        "AVG_WEEKLY_DELIVERY", "DELIVERY_RATIO", "DELIVERY_RATIO_MED",
        "N_HIST_WEEKS", "TRADED_VALUE_CR", "DELIVERY_VALUE_CR",
    ]].copy()

    out["AVG_WEEKLY_DELIVERY"] = out["AVG_WEEKLY_DELIVERY"].round(0)

    out.columns = [
        "SYMBOL", "WEEK_ENDING", "DAYS", "LTP", "PRICE_CHG_%",
        "WEEK_VOLUME", "WEEK_DELIVERY_QTY", "DELIVERY_%",
        "AVG_WEEKLY_DELIVERY", "DELIVERY_RATIO", "DELIVERY_RATIO_MED",
        "HIST_WEEKS", "TRADED_VALUE_CR", "DELIVERY_VALUE_CR",
    ]

    stamp = pd.to_datetime(res["WEEK_ENDING"].iloc[0]).strftime("%Y%m%d")
    outfile = os.path.join(OUTPUT_FOLDER, f"delivery_breakout_weekly_{stamp}.xlsx")
    out.to_excel(outfile, index=False)
    format_excel(outfile)

    print(outfile)
    print(out.head(20).to_string(index=False))
    print(f"\nCandidates: {len(out)} of {len(merged)} symbols scored")
    print(merged["DELIVERY_RATIO"].describe())


def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    for col in ws.columns:
        w = min(max(len(str(x.value)) if x.value else 0 for x in col) + 3, 35)
        ws.column_dimensions[col[0].column_letter].width = w
    wb.save(path)


if __name__ == "__main__":
    st = time.time()
    main()
    print("Execution:", round(time.time() - st, 2), "sec")
